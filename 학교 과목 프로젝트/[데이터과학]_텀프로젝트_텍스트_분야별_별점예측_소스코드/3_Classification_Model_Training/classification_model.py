import openpyxl
import torch
import random
import pickle
import re
import test
#출석체크, 난이도, 학습량, 학점, 성취감
def make_data_set():
    sentences = []
    target_set1 = []
    target_set2 = []
    target_set3 = []
    target_set4 = []
    target_set5 = []
    input_set = []
    hangul = re.compile('[^ ㄱ-ㅣ가-힣]+')

    wb1 = openpyxl.load_workbook('pt_1.xlsx')
    ws1 = wb1.active
    """
    wb2 = openpyxl.load_workbook('pt_2.xlsx')
    ws2 = wb2.active
    """
    wb3 = openpyxl.load_workbook('pt_3.xlsx')
    ws3 = wb3.active
    
    for r1 in ws1.rows:
        if r1[2].value == None:
            break
        sentence = hangul.sub('',r1[2].value)
        for i in sentence:
            if i in test.CHOSUNG_LIST or i in test.JUNGSUNG_LIST or i in test.a:
                sentence = sentence.replace(i, '')
        s_list = sentence.split()
        if not s_list:
            continue
        sentences.append(s_list)
        score_sent = r1[1].value
        target_set1.append(int(score_sent[1]))
        target_set2.append(int(score_sent[4]))
        target_set3.append(int(score_sent[7]))
        target_set4.append(int(score_sent[10]))
        target_set5.append(int(score_sent[13]))

    """
    for r2 in ws2.rows:
        if r2[2].value == None:
            break
        sentence = hangul.sub('', r2[2].value)
        for i in sentence:
            if i in test.CHOSUNG_LIST or i in test.JUNGSUNG_LIST or i in test.a:
                sentence = sentence.replace(i, '')
        s_list = sentence.split()
        if not s_list:
            continue
        sentences.append(s_list)
        score_sent = r2[1].value
        target_set1.append(int(score_sent[1]))
        target_set2.append(int(score_sent[4]))
        target_set3.append(int(score_sent[7]))
        target_set4.append(int(score_sent[10]))
        target_set5.append(int(score_sent[13]))
    """
    for r3 in ws3.rows:
        if r3[2].value == None:
            break
        sentence = hangul.sub('', r3[2].value)
        for i in sentence:
            if i in test.CHOSUNG_LIST or i in test.JUNGSUNG_LIST or i in test.a:
                sentence = sentence.replace(i, '')
        s_list = sentence.split()
        if not s_list:
            continue
        sentences.append(s_list)
        score_sent = r3[1].value
        target_set1.append(int(score_sent[1]))
        target_set2.append(int(score_sent[4]))
        target_set3.append(int(score_sent[7]))
        target_set4.append(int(score_sent[10]))
        target_set5.append(int(score_sent[13]))
    
    with open('embed1\sub2i.txt', 'rb') as s2i:
            sub2i = pickle.load(s2i)
    
    for sentence in sentences:
        subwords = []
        for word in sentence:
            jamo = test.convert(word)
            new_word = list('<' + jamo + '>')
            for n_gram in range(3,7):
                if n_gram > len(new_word):
                    break
                for i in range(len(new_word)):
                    if i >= (len(new_word) - (n_gram-1)):
                        break
                    subword = ''.join(new_word[i:i+n_gram])
                    try:
                        subwords.append(sub2i[subword])
                    except KeyError:
                        pass
        
        input_set.append(subwords)
            
             
    with open('data\input_set.txt', 'wb') as a:
        pickle.dump(input_set, a)
    
    with open('target_set1.txt', 'wb') as b:
        pickle.dump(target_set1, b)

    with open('target_set2.txt', 'wb') as a:
        pickle.dump(target_set2, a)

    with open('target_set3.txt', 'wb') as a:
        pickle.dump(target_set3, a)

    with open('target_set4.txt', 'wb') as a:
        pickle.dump(target_set4, a)

    with open('target_set5.txt', 'wb') as a:
        pickle.dump(target_set5, a)
    
    return

def text_classification(subwords_index, score, input_matrix, output_matrix):
#subwords_index: S
#input_matrix: (V, D)
#output_matrix: (C, D)
#score: (1,World) (2, Sports) (3, Business) (4, Sci/Tech)
    output_matrix = torch.t(output_matrix) #(D, C)

    vec = 0
    for row in input_matrix[subwords_index]:
        vec +=row
    vec = vec / len(subwords_index)
    #projection_layer - output_layer
    W_in = vec #(1, D)
    W_in = torch.reshape(W_in, (1,-1))
    W_out = torch.mm(W_in, output_matrix) #(1, C)
    #softmax_layer
    e = torch.exp(W_out)
    softmax_layer = e / torch.sum(e, dim=1, keepdim=True)

    #predict class
    a = list(softmax_layer[0])
    predict = a.index(max(a))

    #calculate loss
    loss = -torch.log(softmax_layer[0][score])

    softmax_layer[0][score] -= 1
    grad_W_out = softmax_layer #(1, C)

    grad_out = torch.t(torch.mm(torch.t(W_in), grad_W_out)) #(C, D)
    grad_in = torch.mm(grad_W_out, torch.t(output_matrix))

    return loss, grad_in, grad_out, predict

def trainer(target_file, score_kind, learning_rate, epoch):
    print("Making dataset...")
    
    with open('embed1\input_matrix.txt', 'rb') as a:
        input_matrix = pickle.load(a)
    
    with open(target_file, 'rb') as a:
        target_set = pickle.load(a)

    with open('data\input_set.txt', 'rb') as a:
        input_set = pickle.load(a)

    output_matrix = torch.randn(score_kind, input_matrix.shape[1]) / 10  #(C, D)

    print("Training start")
    for k in range(epoch):
        accuracies = []
        loss_counter = 0
        losses = []
        acc = []
        count = 0.01
        for i in range(len(input_set)):
            if i > count*len(input_set):
                print("calculated: %d%% ###############" %(count*100))
                count += 0.01
            
            target = int(target_set[i])
            loss, grad_in, grad_out, predict = text_classification(input_set[i], target-1, input_matrix, output_matrix)
            output_matrix -= learning_rate*grad_out

            losses.append(loss)
            loss_counter += 1

            if target == predict + 1:
                accuracies.append(1)
                acc.append(1)
            else:
                accuracies.append(0)
                acc.append(0)

            if loss_counter % 1000 == 0:
                avg_loss = sum(losses) / len(losses)
                print("Loss: %f , Accuracy: %f%%\n" %(avg_loss, sum(accuracies) / len(accuracies) * 100))
                losses = []
                accuracies = []
    print("total acc: " + sum(acc)/len(acc) )
    with open('data\output_matrix5.txt', 'wb') as OM:
        pickle.dump(output_matrix, OM)
    
    

def tester(file_name, n):
    with open('input_matrix.txt', 'rb') as IM:
        input_matrix = pickle.load(IM)

    with open('output_matrix.txt', 'rb') as OM:
        output_matrix = pickle.load(OM)
    
    with open('n_gram_dict.txt', 'rb') as DICT:
        trained_ngram_dict = pickle.load(DICT)

    print("Making dataset...")
    input_set, target_set, sentence_dict = make_dataset_TEST(file_name, n, trained_ngram_dict)

    accuracies = []
    acc = []
    loss_counter = 0
    losses = []
    count = 0.01

    print("Testing start")
    for i in range(len(input_set)):
        if i > count*len(input_set):
            print("calculated: %f%% #######################" %(count*100))
            count += 0.01
        target = int(target_set[i]) 
        if not input_set[i]:
            continue        
        else:                        
            loss, _, _, predict = text_classification(input_set[i], target - 1, input_matrix, output_matrix)

            losses.append(loss)
            loss_counter += 1

            if target == predict + 1:
                accuracies.append(1)
                acc.append(1)
            else:
                accuracies.append(0)
                acc.append(0)
            if loss_counter % 1000 == 0:
                avg_loss = sum(losses) / len(losses)
                print("Loss: %f , Accuracy: %f%%\n" %(avg_loss, sum(accuracies) / len(accuracies) * 100))
                losses = []
                accuracies = []
        print("total accuracy:" + sum(acc)/len(acc))


trainer('target_set5.txt', 5, 0.025, 1)

#tester('test.csv', 2)