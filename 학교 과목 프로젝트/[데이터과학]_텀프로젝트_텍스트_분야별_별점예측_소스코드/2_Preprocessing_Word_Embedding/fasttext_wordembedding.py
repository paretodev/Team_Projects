import torch
import random
from collections import Counter
import operator
import pickle
import timeit
import openpyxl
import test
import re
"""
class NN(word_dimension, embedded_dimesion):
    def __init(self, word_dimension, embedded_dimension):
        self.embed_layer = torch.Tensor(word_dimension, embedded_dimension)
        self.W_out = torch.Tensor(embedded_demension, word_dimension)
"""
def make_corpus(file_name, mode): 
    # have to make capital case >> lower case
    """
    f = open(file_name, 'r')

    String = f.read()
    if mode == 'part':
        String = String[0:int(len(String)/50)]
    elif mode == 'full':
        pass

    String = String.lower()

    whole_corpus = String.split()
    freq_table = Counter(whole_corpus)
    
    corpus = []
    for word in whole_corpus:
        if freq_table[word] > 4:
            corpus.append(word)
    
    f.close
    """
    corpus = []
    corpus1 = []
    corpus2 = []
    corpus3 = []
    sentences = []
    hangul = re.compile('[^ ㄱ-ㅣ가-힣]+')

    wb1 = openpyxl.load_workbook('pt_1.xlsx')
    ws1 = wb1.active
    
    wb2 = openpyxl.load_workbook('pt_2.xlsx')
    ws2 = wb2.active
    wb3 = openpyxl.load_workbook('pt_3.xlsx')
    ws3 = wb3.active
    
    for r1 in ws1.rows:
        if r1[2].value == None:
            break
        sentence = hangul.sub('',r1[2].value)
        for i in sentence:
            if i in test.CHOSUNG_LIST or i in test.JUNGSUNG_LIST or i in test.a:
                sentence = sentence.replace(i, '')
        s_list = sentence.split()
        if not s_list:
            continue
        corpus1 += s_list
        sentences.append(s_list)
    
    for r2 in ws2.rows:
        if r2[2].value == None:
            break
        sentence = hangul.sub('', r2[2].value)
        for i in sentence:
            if i in test.CHOSUNG_LIST or i in test.JUNGSUNG_LIST or i in test.a:
                sentence = sentence.replace(i, '')
        s_list = sentence.split()
        if not s_list:
            continue
        corpus2 += s_list
        sentences.append(s_list)

    for r3 in ws3.rows:
        if r3[2].value == None:
            break
        sentence = hangul.sub('', r3[2].value)
        for i in sentence:
            if i in test.CHOSUNG_LIST or i in test.JUNGSUNG_LIST or i in test.a:
                sentence = sentence.replace(i, '')
        s_list = sentence.split()
        if not s_list:
            continue
        corpus3 += s_list
        sentences.append(s_list)
    
    corpus = corpus1 + corpus2 + corpus3
    return corpus, sentences  # word sequence list


def make_word_list(corpus):
    """
    word_index_list = list(range(len(corpus)))
    word_list = {}
    for word in corpus:
        index = random.choice(word_index_list)
        word_list[word] = word_index_list.pop(index)
    """
    
    freq_table = Counter(corpus)

    word_list = freq_table.keys()
    
    return word_list, freq_table # kinds of word list, frequency table for negative sampling

def make_word_dict(word_list):
    word_dict = {} #word - index dictionary
    """
    for word in word_list:
        if word not in word_dict:
            word_dict[word] = 0
        
    word_dict_length = list(range(len(word_dict.keys())))

    for word in word_dict:
        index = random.choice(word_dict_length)
        word_dict[word] = index
        word_dict_length.remove(index)
    """
    word_list = list(set(word_list))
    counter = 0
    for word in word_list:
        word_dict[word] = counter
        counter = counter + 1
    
    return word_dict


def make_subword_dict(word_list, n_gram_min, n_gram_max): 
    subword_dict = {} # word - subword dictionary
    sub2i = {} # subwords to index
    subword_list = []
    for word in word_list:
        subwords = []
        jamo = test.convert(word)
        new_word = list('<' + jamo + '>')
        for n_gram in range(n_gram_min, n_gram_max + 1): #make n-gram subword token
            if n_gram > len(new_word):
                break
            for i in range(len(new_word)):
                if i >= (len(new_word) - (n_gram-1)):
                    break

                subword = ''.join(new_word[i:i+n_gram])
                subwords.append(subword)
                subword_list.append(subword)

        subword_dict[word] = subwords
    
    subword_list = list(set(subword_list))

    counter = 0
    for subword in subword_list:
        sub2i[subword] = counter
        counter += 1

    """
    for word in word_list:
        subwords = []
        new_word = list('<' + word + '>') #blank space describe by '<' , '>'
        for n_gram in range(n_gram_min, n_gram_max + 1): #make n-gram subword token
            if n_gram > len(new_word):
                break
            for i in range(len(new_word)):
                if i >= (len(new_word) - (n_gram-1)):
                    break

                subword = ''.join(new_word[i:i+n_gram])
                subwords.append(subword)

                if subword not in sub2i.keys():
                    sub2i[subword] = 0
        
        special_token = ''.join(new_word)
        if special_token not in subwords:
            subwords.append(special_token)
            sub2i[special_token] = 0
        
        subword_dict[word] = subwords


    sub2i_length = list(range(len(sub2i)))
    for subword in sub2i.keys():
        index = random.choice(sub2i_length)
        sub2i[subword] = index
        sub2i_length.remove(index)
    """
    return subword_dict, sub2i


def subsampling(corpus):
###############################  Output  #########################################
# subsampled : Subsampled corpus                                             #
##################################################################################
    num_word = Counter(corpus)
    word_count = len(corpus)

    for word in num_word:
        num_word[word] = 1 - (0.00001 / (num_word[word] / word_count))**0.5
    
    word_prob = num_word

    subsampled = []
    for word in corpus:
        if random.uniform(0, 1) > word_prob[word]:
            subsampled.append(word)
    
    return subsampled

def make_dataset(subsampled_sentences, word_dict, subword_dict, sub2i, window_size):
    input_set = [] #subwords indices per 1word is element of list
    target_set = [] #indices of contextwords per 1word is element of list

    for subsampled in subsampled_sentences:
        subsampled_length = len(subsampled)
        if subsampled_length < window_size*2 + 1:
            continue
        for j in range(subsampled_length):
            word = subsampled[j]
            subwords = subword_dict[word]
            sub_indices = [sub2i[subword] for subword in subwords]
            input_set.append(sub_indices)

            if j < window_size:
                context_indices = [word_dict[subsampled[i]] for i in range(j)] + [word_dict[subsampled[i]] for i in range(j+1, j+1+window_size)]
                target_set.append(context_indices)
            elif j > len(subsampled) - window_size -1:
                context_indices = [word_dict[subsampled[i]] for i in range(j-1-window_size, j-1)] + [word_dict[subsampled[i]] for i in range(j+1, subsampled_length)]
                target_set.append(context_indices)
            else:
                context_indices = [word_dict[subsampled[i]] for i in range(j-1-window_size, j-1)] + [word_dict[subsampled[i]] for i in range(j+1, j+1+window_size)]
                target_set.append(context_indices)

    return input_set, target_set


def fasttext(sub_indices, context_word, input_matrix, output_matrix):
    #sub_indices: indices of subwords of center word <int list>
    #context_word: index of contextword <int>
    #input_matrix: (S, D)
    #output_matrix: (V, D)
    #S: dimension of subwords (the number of subwords)
    #D: dimension of embedding vector
    #V: 

    # subwords -- <input matrix> -- subwords_layer -- <sum> -- embed layer -- <output matrix> -- output layer -- <soft max> -- softmax layer -- <cross entropy> -- loss
    #<>: operation
    output_matrix = torch.t(output_matrix) #(V, D)
    subwords_layer = input_matrix[sub_indices] #(M, D)
    embed_layer = 0

    for row in subwords_layer:
        embed_layer += row


    embed_layer = torch.reshape(embed_layer, (1,-1)) #embed_layer: (1, D)
    output_layer = torch.mm(embed_layer, output_matrix) #output_layer: (1, V)

    e = torch.exp(output_layer)
    softmax_layer = e / torch.sum(e, dim=1, keepdim=True)

    loss = -torch.log(softmax_layer[0][context_word])

    softmax_layer[0][context_word] -= 1
    grad_output_layer = softmax_layer #(1, V)

    grad_out = torch.mm(torch.t(embed_layer), grad_output_layer) #(D, V)

    grad_out = torch.t(grad_out)

    grad_embed_layer = torch.mm(grad_output_layer, torch.t(output_matrix)) #(1, D)
    grad_sum_layer = grad_embed_layer #(1, D)
    """
    grad_subwords_layer = torch.zeros((len(sub_indices), input_matrix.shape[1])) #(M, D)
    for i in grad_subwords_layer:
        i = grad_sum_layer
    """
    grad_in = grad_sum_layer #(1, D)
    
    return loss, grad_in, grad_out

def negative_sampling(freq, target_set):
    sampling = []

    for i in range(15):
        choice = random.choice(freq)

        if choice == target_set:
            i = i-1
        else:
            sampling.append(int(choice))

    sampling.append(target_set)
    random.shuffle(sampling)
    index = sampling.index(target_set)
    return sampling, index


def cos_similarity(target_word_list, word_dict, subword_dict, sub2i, input_matrix, n_gram_min, n_gram_max):
    word_vec_dict = {} #word - vector dictionary
    cos_similar_dict = {} #word - cosine similarity with target word dictionary
    for i in word_dict:
            sub_index_list = [] 
            word_vectors = 0
            for subword in subword_dict[i]:
                sub_index_list.append(sub2i[subword])
            
            for row in input_matrix[sub_index_list]:
                word_vectors += row

            word_vec_dict[i] = word_vectors

    for target_word in target_word_list:
        subwords = [] #target word's subword list
        new_word = list('<' + target_word + '>') #blank space describe by '<' , '>'
        for n_gram in range(n_gram_min, n_gram_max + 1): #make n-gram subword token
            if n_gram > len(new_word):
                break
            for i in range(len(new_word)):
                if i >= (len(new_word) - (n_gram-1)):
                    break

                subword = ''.join(new_word[i:i+n_gram])
                if subword in sub2i:
                    subwords.append(subword)
        
        special_token = ''.join(new_word)
        if special_token not in subwords:
            if special_token in sub2i:
                subwords.append(special_token)
        
        sub_indices = [sub2i[subword] for subword in subwords] 
        sub_vec = input_matrix[sub_indices] #vectors of subwords of target word

        word_vec = 0 #vector of target word
        for row in sub_vec:
            word_vec += row
        
        word_vec = torch.Tensor(word_vec)
        for i in word_vec_dict.keys():
            word_vec = torch.reshape(word_vec, (1,-1))
            word_vec_dict_reshape = torch.reshape(word_vec_dict[i], (1,-1))
            cos_similar_dict[i] = torch.cosine_similarity(word_vec, word_vec_dict_reshape)
        
        sorted_cos = sorted(cos_similar_dict.items(), key=lambda x: x[1], reverse = True)

        print('word: ' + str(target_word))
        for i in range(1,6):
            print(str(sorted_cos[i]))
        print('\n')

    return


def main(file_name, window_size, n_gram_min, n_gram_max, dimension, learning_rate, mode, epoch, CBOW):
    print("Read file...\n")
    corpus, sentences = make_corpus(file_name, mode)
    
    print("Subsampling...\n")
    #corpus = subsampling(corpus) 
    #if corpus is too short, do not subsample

    print("Making vocab...\n")
    word_list, freq_table = make_word_list(corpus)


    print("Making dictionary...\n")
    word_dict = make_word_dict(word_list)

    freq = []
    for k, v in freq_table.items():
        f = int(v**0.75)
        for _ in range(f):
            freq.append(word_dict[k])
        
    print("Making subword...\n")
    subword_dict, sub2i = make_subword_dict(word_list, n_gram_min, n_gram_max)

    with open('word_dict.txt', 'wb') as a:
        pickle.dump(word_dict, a)

    with open('subword_dict.txt', 'wb') as b:
        pickle.dump(subword_dict, b)

    with open('sub2i.txt', 'wb') as c:
        pickle.dump(sub2i, c)
    
    print("Making training set...\n")
    if CBOW == 'CBOW':
        input_set, target_set = make_dataset_CBOW(sentences, word_dict, subword_dict, sub2i, window_size = window_size)
    else:
        input_set, target_set = make_dataset(sentences, word_dict, subword_dict, sub2i, window_size = window_size)



    loss_counter = 0
    losses = []
    avg_loss = None
    print("Train start\n")
    if CBOW == 'CBOW':
        input_matrix = torch.randn(len(sub2i), dimension)  / 20 #/ (dimension)**0.5
        output_matrix = torch.randn(dimension, len(word_list)) / 20  #/ (dimension)**0.5
        for k in range(epoch):
            second_break = [True]
            count = 0.01
            start = timeit.default_timer()

            for i in range(len(input_set)):

                if i > count*len(input_set):
                    print("calculated: %d%%" %(count*100))
                    count += 0.01

                activated, index = negative_sampling(freq, target_set[i])
                loss, grad_in, grad_out = fasttext_CBOW_NS(input_set[i], index, input_matrix, output_matrix[:, [activated]])
                for vector in input_set[i]:
                    input_matrix[vector] -= learning_rate*grad_in
                
                output_matrix = torch.t(output_matrix)
                output_matrix[activated] -= learning_rate*torch.t(grad_out)
                output_matrix = torch.t(output_matrix)
                #output_matrix[activated] -= learning_rate*grad_out
                losses.append(loss)    

                #print("loss = %f" %(loss))
                if loss_counter % 1000 == 0:
                    avg_loss = sum(losses) / len(losses)
                    print("Loss: %f \n" %(avg_loss))
                    losses = []
                loss_counter += 1

                end = timeit.default_timer()

                if end - start > 3600*12:
                    print('time over!!')
                    break
    else:
        input_matrix = torch.randn(len(sub2i), dimension)  / 10 #/ (dimension)**0.5
        output_matrix = torch.randn(dimension, len(word_list)) / 10  #/ (dimension)**0.5

        for k in range(epoch):
            second_break = [True]
            count = 0.01
            start = timeit.default_timer()
            for i in range(len(input_set)):

                if i > count*len(input_set):
                    print("calculated: %d%%" %(count*100))
                    count += 0.01

                for j in target_set[i]:
                    output_matrix = torch.t(output_matrix)
                    activated, index = negative_sampling(freq, j)
                    loss, grad_in, grad_out = fasttext(input_set[i], index, input_matrix, output_matrix[activated])
                    input_matrix[input_set[i]] -= learning_rate*grad_in
                    output_matrix[activated] -= learning_rate*grad_out
                    output_matrix = torch.t(output_matrix)
                    losses.append(loss)    

                    #print("loss = %f" %(loss))
                    if loss_counter % 1000 == 0:
                        avg_loss = sum(losses) / len(losses)
                        print("Loss: %f \n" %(avg_loss))
                        losses = []
                    loss_counter += 1

                end = timeit.default_timer()

                if end - start > 3600*12:
                    print('time over!!')
                    break
                

    with open('input_matrix.txt ', 'wb') as f:
        pickle.dump(input_matrix, f)


    
    #target_word_list = ['출석', '성적', '빡세', '시험', '열심']
    #target_word_list = ['anarc', 'wor', 'he']
    #cos_similarity(target_word_list, word_dict, subword_dict, sub2i, input_matrix, n_gram_min, n_gram_max)
    return

main('text8.txt', 2, 3, 6, 200, 0.01, 'full', 1, 'CBOsW')
